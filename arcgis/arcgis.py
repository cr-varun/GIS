from typing import Iterator, Tuple

import laspy
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os
import csv
import sys


EXCEL_MAX_ROWS = 1_048_576  # per sheet including header
DEFAULT_LAS = "./sample.las"
CHUNK_SIZE = 500_000


def chunk_points(reader: laspy.LasReader, chunk_size: int) -> Iterator[Tuple]:
    """Yield tuples of (x, y, z) arrays in chunks."""
    # laspy 2.x provides reader.chunk_iterator
    for points in reader.chunk_iterator(chunk_size):
        # Use scaled coordinates (float) via .x/.y/.z
        yield points.x, points.y, points.z


def export_las_to_excel(las_path: str, xlsx_path: str, chunk_size: int = 500_000) -> None:
    """
    Stream easting (X), northing (Y), elevation (Z) from a .las/.laz to an Excel file.

    - Automatically splits across multiple sheets if row limit is exceeded.
    - Uses scaled coordinates so units match the LAS header.
    """
    wb = Workbook(write_only=True)
    sheet_index = 1
    ws = wb.create_sheet(title=f"Points_{sheet_index}")
    # Header
    ws.append(["EASTING", "NORTHING", "ELEVATION"])  # X, Y, Z
    rows_in_sheet = 1  # header counted

    with laspy.open(las_path) as reader:
        for x_arr, y_arr, z_arr in chunk_points(reader, max(1, chunk_size)):
            # Prepare rows for this chunk
            for row in zip(x_arr, y_arr, z_arr):
                if rows_in_sheet >= EXCEL_MAX_ROWS:
                    # new sheet
                    sheet_index += 1
                    ws = wb.create_sheet(title=f"Points_{sheet_index}")
                    ws.append(["EASTING", "NORTHING", "ELEVATION"])  # header
                    rows_in_sheet = 1
                ws.append(row)
                rows_in_sheet += 1

    # Remove default sheet if present and empty (write_only mode sometimes adds one)
    # Ensure the first sheet is the first created sheet
    # Save workbook
    wb.save(xlsx_path)
    try:
        wb.close()
    except Exception:
        pass


def export_las_to_csv(las_path: str, csv_path: str, chunk_size: int = 1_000_000) -> None:
    """Fast streaming CSV export for large point clouds."""
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["EASTING", "NORTHING", "ELEVATION"])  # header
        with laspy.open(las_path) as reader:
            for x_arr, y_arr, z_arr in chunk_points(reader, max(1, chunk_size)):
                writer.writerows(zip(x_arr, y_arr, z_arr))


def main():
    # Optional headless mode: set env LAS_EXPORT_HEADLESS=1 to export DEFAULT_LAS quickly without GUI
    if os.environ.get("LAS_EXPORT_HEADLESS") == "1":
        las_path = DEFAULT_LAS
        out_path = las_path.rsplit(".", 1)[0] + ".csv"
        export_las_to_csv(las_path, out_path, CHUNK_SIZE)
        print(f"Headless export complete: {out_path}")
        return

    # Simple non-GUI mode: if file paths are provided as arguments, export directly (CSV by default for speed)
    if len(sys.argv) > 1:
        files = [p for p in sys.argv[1:] if os.path.isfile(p)]
        if not files:
            print("No valid input files provided.")
            return
        for las_path in files:
            base = os.path.splitext(os.path.basename(las_path))[0]
            out_dir = os.path.dirname(las_path)
            out_csv = os.path.join(out_dir, base + ".csv")
            export_las_to_csv(las_path, out_csv, CHUNK_SIZE)
            print(f"Exported: {out_csv}")
        return

    class App(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("LAS to Excel Exporter")
            self.geometry("700x450")
            self.selected_files = []
            self.output_dir = ""
            self.as_csv_var = tk.BooleanVar(value=True)  # CSV is faster

            btn_frame = tk.Frame(self)
            btn_frame.pack(fill=tk.X, padx=10, pady=10)

            self.import_btn = tk.Button(btn_frame, text="Import (.las/.laz)", command=self.on_import)
            self.import_btn.pack(side=tk.LEFT, padx=5)

            self.out_btn = tk.Button(btn_frame, text="Output Folder", command=self.on_choose_output)
            self.out_btn.pack(side=tk.LEFT, padx=5)

            self.export_btn = tk.Button(btn_frame, text="Export to Excel", command=self.on_export)
            self.export_btn.pack(side=tk.LEFT, padx=5)

            self.csv_chk = tk.Checkbutton(btn_frame, text="Export as CSV (faster)", variable=self.as_csv_var)
            self.csv_chk.pack(side=tk.LEFT, padx=10)

            out_frame = tk.Frame(self)
            out_frame.pack(fill=tk.X, padx=10)
            tk.Label(out_frame, text="Output:").pack(side=tk.LEFT)
            self.out_var = tk.StringVar(value="(same folder as input)")
            self.out_entry = tk.Entry(out_frame, textvariable=self.out_var, state="readonly")
            self.out_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

            list_frame = tk.Frame(self)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            self.listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED)
            self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar = tk.Scrollbar(list_frame, command=self.listbox.yview)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.listbox.config(yscrollcommand=scrollbar.set)

            prog_frame = tk.Frame(self)
            prog_frame.pack(fill=tk.X, padx=10, pady=(0,10))
            self.progress = ttk.Progressbar(prog_frame, orient=tk.HORIZONTAL, mode="determinate")
            self.progress.pack(fill=tk.X)
            self.status_var = tk.StringVar(value="Ready")
            self.status = tk.Label(self, textvariable=self.status_var, anchor="w")
            self.status.pack(fill=tk.X, padx=10, pady=(0,10))

        def on_import(self):
            paths = filedialog.askopenfilenames(
                title="Select LAS/LAZ files",
                filetypes=[("LAS/LAZ Files", "*.las *.laz"), ("All Files", "*.*")],
            )
            if not paths:
                return
            self.selected_files = list(paths)
            self.listbox.delete(0, tk.END)
            for p in self.selected_files:
                self.listbox.insert(tk.END, p)
            self.status_var.set(f"Selected {len(self.selected_files)} file(s)")

        def on_choose_output(self):
            d = filedialog.askdirectory(title="Select output folder")
            if not d:
                return
            self.output_dir = d
            self.out_var.set(self.output_dir)

        def on_export(self):
            if not self.selected_files:
                messagebox.showinfo("No files", "Import .las/.laz files first")
                return
            self.disable_controls()
            self.progress.configure(maximum=len(self.selected_files), value=0)
            self.status_var.set("Exporting...")
            threading.Thread(target=self._export_worker, daemon=True).start()

        def _export_worker(self):
            ok = 0
            for idx, las_path in enumerate(self.selected_files, start=1):
                try:
                    base = os.path.splitext(os.path.basename(las_path))[0]
                    if self.output_dir:
                        x_path = os.path.join(self.output_dir, base + (".csv" if self.as_csv_var.get() else ".xlsx"))
                    else:
                        x_path = os.path.join(os.path.dirname(las_path), base + (".csv" if self.as_csv_var.get() else ".xlsx"))
                    if self.as_csv_var.get():
                        export_las_to_csv(las_path, x_path, CHUNK_SIZE)
                    else:
                        export_las_to_excel(las_path, x_path, CHUNK_SIZE)
                    ok += 1
                except Exception as e:
                    self.after(0, lambda p=las_path, msg=str(e): messagebox.showerror("Error", f"{p}\n{msg}"))
                finally:
                    self.after(0, lambda v=idx: self.progress.configure(value=v))
            self.after(0, lambda n=ok: self._export_done(n))

        def _export_done(self, ok):
            self.status_var.set(f"Completed: {ok}/{len(self.selected_files)}")
            self.enable_controls()
            if ok > 0:
                messagebox.showinfo("Done", f"Exported {ok} file(s)")

        def disable_controls(self):
            for w in (self.import_btn, self.out_btn, self.export_btn):
                w.configure(state=tk.DISABLED)

        def enable_controls(self):
            for w in (self.import_btn, self.out_btn, self.export_btn):
                w.configure(state=tk.NORMAL)

    App().mainloop()


if __name__ == "__main__":
    main()
